from __future__ import annotations

from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

from .formatting import fmt_usd
from .models import ReportData, TokenFlow


TOKEN_DECIMALS = {
    "BNB": 3,
    "ETH": 3,
    "DAI": 0,
    "USDC": 0,
    "USDT": 0,
}
NAVY = "102A43"
GREEN = "166534"
RED = "991B1B"
DARK = "1F2937"
MUTED = "64748B"
GRID = "E2E8F0"
SUMMARY_BG = "F5F8FB"
HEADER_BG = "EAF1F7"
GREEN_BG = "DCFCE7"
RED_BG = "FEE2E2"
BOTTOM_BG = "E9F7EF"
WHITE = "FFFFFF"
COMPAT_FONT = "Arial"


def render_report_files(report: ReportData, output_dir: Path, history_reports: Sequence[ReportData] | None = None) -> tuple[Path, Path]:
    run_dir = output_dir / report.captured_at.strftime("%Y%m%d")
    run_dir.mkdir(parents=True, exist_ok=True)
    base = "anubis_report_" + report.captured_at.strftime("%Y%m%d_%H%M%S")
    image_path = run_dir / f"{base}.png"
    xlsx_path = run_dir / f"{base}.xlsx"
    render_png(report, image_path)
    render_xlsx(report, xlsx_path, history_reports=history_reports)
    return image_path, xlsx_path


def render_xlsx(report: ReportData, path: Path, history_reports: Sequence[ReportData] | None = None) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = f"{report.captured_at.month}月{report.captured_at.day}汇总"
    _render_example_sheet(ws, report)

    history_ws = wb.create_sheet("歷史記錄")
    _render_history_sheet(history_ws, history_reports or ())
    wb.active = 0
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _render_example_sheet(ws, report: ReportData, start_row: int = 1, freeze_panes: bool = True) -> None:
    ws.sheet_view.showGridLines = False

    row = lambda value: start_row + value - 1
    _merge_example_ranges(ws, len(report.tokens), start_row=start_row)
    ws.cell(row(1), 1, f"币种资金流日报 - {report.captured_at.strftime('%Y-%m-%d')}")
    ws.cell(row(2), 1, "转入美金合计")
    ws.cell(row(2), 3, "转出美金合计")
    ws.cell(row(2), 5, report.status)
    ws.cell(row(2), 7, "记录日期")
    ws.cell(row(3), 1, float(_money4(report.total_in_usd)))
    ws.cell(row(3), 3, float(_money4(report.total_out_usd)))
    ws.cell(row(3), 5, float(_money4(report.net_usd)))
    ws.cell(row(3), 7, report.captured_at.strftime("%Y-%m-%d %H:%M:%S HKT"))
    ws.cell(row(5), 1, "币种")
    ws.cell(row(5), 2, "转入数量")
    ws.cell(row(5), 4, "转出数量")
    ws.cell(row(5), 6, "净数量")
    ws.cell(row(5), 8, "状态")

    for row_offset, token in enumerate(report.tokens, row(6)):
        decimals = _token_decimals(token.symbol)
        net_qty = _round_token_qty(token.net_qty, token.symbol)
        ws.cell(row_offset, 1, token.symbol)
        ws.cell(row_offset, 2, float(_round_token_qty(token.in_qty, token.symbol)))
        ws.cell(row_offset, 4, float(_round_token_qty(token.out_qty, token.symbol)))
        ws.cell(row_offset, 6, float(net_qty))
        ws.cell(row_offset, 8, _status_label(net_qty))
        number_format = _qty_number_format(decimals)
        ws.cell(row_offset, 2).number_format = number_format
        ws.cell(row_offset, 4).number_format = number_format
        ws.cell(row_offset, 6).number_format = number_format

    _style_example_workbook(ws, report, start_row=start_row, freeze_panes=freeze_panes)


def _render_history_sheet(ws, reports: Sequence[ReportData]) -> None:
    ws.sheet_view.showGridLines = False
    if not reports:
        ws["A1"] = "歷史記錄"
        ws["A2"] = "尚未有每日 23:30 HKT 自動報表記錄。"
        ws["A1"].font = Font(name=COMPAT_FONT, size=18, bold=True, color=DARK)
        ws["A2"].font = Font(name=COMPAT_FONT, size=12, color=MUTED)
        ws.column_dimensions["A"].width = 46
        return

    start_row = 1
    for report in reports:
        _render_example_sheet(ws, report, start_row=start_row, freeze_panes=start_row == 1)
        start_row += _example_block_height(report)


def _render_native_compatible_sheet(ws, report: ReportData) -> None:
    """First sheet uses only plain cells for Apple Quick Look, Numbers, and iOS preview."""
    ws.sheet_view.showGridLines = True
    ws.freeze_panes = None

    title = f"币种资金流日报 - {report.captured_at.strftime('%Y-%m-%d')}"
    rows = [
        ["项目", "内容", ""],
        ["报表标题", title, ""],
        ["记录日期", report.captured_at.strftime("%Y-%m-%d %H:%M:%S HKT"), ""],
        ["转入美金合计", _fmt_usd4(report.total_in_usd), ""],
        ["转出美金合计", _fmt_usd4(report.total_out_usd), ""],
        [report.status, _fmt_usd4(report.net_usd), ""],
        ["", "", ""],
        ["币种", "转入数量", "转出数量", "净数量", "状态"],
    ]
    for token in report.tokens:
        net_qty = _round_token_qty(token.net_qty, token.symbol)
        rows.append(
            [
                token.symbol,
                _fmt_token_qty(token.in_qty, token.symbol),
                _fmt_token_qty(token.out_qty, token.symbol),
                _fmt_token_qty(net_qty, token.symbol),
                _status_label(net_qty),
            ]
        )

    for row_idx, row_values in enumerate(rows, 1):
        for col_idx, value in enumerate(row_values, 1):
            ws.cell(row_idx, col_idx, value)

    widths = [12, 18, 18, 18, 10]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    thin = Side(style="thin", color=GRID)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor=HEADER_BG)
    summary_fill = PatternFill("solid", fgColor=SUMMARY_BG)
    net_fill = PatternFill("solid", fgColor=GREEN_BG if report.net_usd > 0 else RED_BG if report.net_usd < 0 else SUMMARY_BG)
    net_color = GREEN if report.net_usd > 0 else RED if report.net_usd < 0 else DARK

    for row in ws.iter_rows(min_row=1, max_row=len(rows), min_col=1, max_col=5):
        for cell in row:
            cell.font = Font(name=COMPAT_FONT, size=11, color=DARK)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            cell.number_format = "@"

    for row_idx in range(1, 7):
        for col_idx in range(1, 6):
            ws.cell(row_idx, col_idx).fill = summary_fill
    for col_idx in range(1, 6):
        ws.cell(1, col_idx).fill = PatternFill("solid", fgColor=NAVY)
        ws.cell(1, col_idx).font = Font(name=COMPAT_FONT, size=11, bold=True, color=WHITE)
        ws.cell(8, col_idx).fill = header_fill
        ws.cell(8, col_idx).font = Font(name=COMPAT_FONT, size=11, bold=True, color=DARK)

    ws["B2"].font = Font(name=COMPAT_FONT, size=13, bold=True, color=DARK)
    for cell_ref in ["B2", "B3", "B4", "B5", "B6"]:
        ws[cell_ref].alignment = Alignment(horizontal="left", vertical="center")
    for cell_ref in ["A6", "B6", "C6", "D6", "E6"]:
        ws[cell_ref].fill = net_fill
        ws[cell_ref].font = Font(name=COMPAT_FONT, size=11, bold=True, color=net_color)

    for row_idx in range(9, 9 + len(report.tokens)):
        net_text = str(ws.cell(row_idx, 4).value or "")
        fill = PatternFill("solid", fgColor=WHITE)
        color = DARK
        if net_text.startswith("-"):
            fill = PatternFill("solid", fgColor=RED_BG)
            color = RED
        elif net_text not in {"0", "0.0", "0.000", ""}:
            fill = PatternFill("solid", fgColor=GREEN_BG)
            color = GREEN
        ws.cell(row_idx, 1).font = Font(name=COMPAT_FONT, size=11, bold=True, color=DARK)
        ws.cell(row_idx, 4).fill = fill
        ws.cell(row_idx, 4).font = Font(name=COMPAT_FONT, size=11, bold=True, color=color)

    for row_idx in range(1, len(rows) + 1):
        ws.row_dimensions[row_idx].height = 24


def workbook_has_formulas(path: Path) -> bool:
    wb = load_workbook(path, data_only=False)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    return True
    return False


def render_png(report: ReportData, path: Path) -> None:
    rows = len(report.tokens)
    col_w = [128, 160, 180, 160, 180, 150, 200, 110]
    row_h = [54, 24, 39, 34] + [34] * rows + [14]
    width = sum(col_w) + 2
    height = sum(row_h) + 2

    img = Image.new("RGB", (width, height), "#" + WHITE)
    draw = ImageDraw.Draw(img)
    fonts = {
        "title": _font(36, bold=True),
        "summary_label": _font(15, bold=True),
        "summary_value": _font(18, bold=True),
        "header": _font(15, bold=True),
        "symbol": _font(14, bold=True),
        "number": _font(23),
        "net": _font(23, bold=True),
        "status": _font(14),
    }

    xs = [1]
    for width_item in col_w:
        xs.append(xs[-1] + width_item)
    ys = [1]
    for height_item in row_h:
        ys.append(ys[-1] + height_item)

    _draw_cell(draw, xs, ys, 0, 0, 7, 0, NAVY)
    _draw_center(draw, (xs[0], ys[0], xs[8], ys[1]), f"币种资金流日报 - {report.captured_at.strftime('%Y-%m-%d')}", fonts["title"], WHITE, y_offset=-1)

    summary_fill = GREEN_BG if report.net_usd > 0 else RED_BG if report.net_usd < 0 else SUMMARY_BG
    summary_color = GREEN if report.net_usd > 0 else RED if report.net_usd < 0 else DARK
    for start_col, end_col, label, value, fill, value_color in [
        (0, 1, "转入美金合计", _fmt_usd4(report.total_in_usd), SUMMARY_BG, DARK),
        (2, 3, "转出美金合计", _fmt_usd4(report.total_out_usd), SUMMARY_BG, DARK),
        (4, 5, report.status, _fmt_usd4(report.net_usd), summary_fill, summary_color),
        (6, 7, "记录日期", report.captured_at.strftime("%Y-%m-%d %H:%M:%S HKT"), SUMMARY_BG, DARK),
    ]:
        _draw_cell(draw, xs, ys, start_col, 1, end_col, 1, fill)
        _draw_cell(draw, xs, ys, start_col, 2, end_col, 2, fill)
        _draw_center(draw, (xs[start_col], ys[1], xs[end_col + 1], ys[2]), label, fonts["summary_label"], MUTED if fill == SUMMARY_BG else summary_color)
        _draw_center(draw, (xs[start_col], ys[2], xs[end_col + 1], ys[3]), value, fonts["summary_value"], value_color)

    for start_col, end_col, label in [
        (0, 0, "币种"),
        (1, 2, "转入数量"),
        (3, 4, "转出数量"),
        (5, 6, "净数量"),
        (7, 7, "状态"),
    ]:
        _draw_cell(draw, xs, ys, start_col, 3, end_col, 3, HEADER_BG)
        _draw_center(draw, (xs[start_col], ys[3], xs[end_col + 1], ys[4]), label, fonts["header"], DARK)

    for idx, token in enumerate(report.tokens):
        row_idx = 4 + idx
        net_qty = _round_token_qty(token.net_qty, token.symbol)
        net_fill = GREEN_BG if net_qty > 0 else RED_BG if net_qty < 0 else WHITE
        net_color = GREEN if net_qty > 0 else RED if net_qty < 0 else DARK
        row_data = [
            (0, 0, token.symbol, fonts["symbol"], DARK, WHITE),
            (1, 2, _fmt_token_qty(token.in_qty, token.symbol), fonts["number"], "111827", WHITE),
            (3, 4, _fmt_token_qty(token.out_qty, token.symbol), fonts["number"], "111827", WHITE),
            (5, 6, _fmt_token_qty(net_qty, token.symbol), fonts["net"], net_color, net_fill),
            (7, 7, _status_label(net_qty), fonts["status"], DARK, WHITE),
        ]
        for start_col, end_col, text, font, color, fill in row_data:
            _draw_cell(draw, xs, ys, start_col, row_idx, end_col, row_idx, fill)
            _draw_center(draw, (xs[start_col], ys[row_idx], xs[end_col + 1], ys[row_idx + 1]), text, font, color)

    bottom_row = 4 + rows
    _draw_cell(draw, xs, ys, 0, bottom_row, 7, bottom_row, BOTTOM_BG)
    path.parent.mkdir(parents=True, exist_ok=True)
    img.save(path)


def _status_color(value: Decimal) -> str:
    if value > 0:
        return GREEN
    if value < 0:
        return RED
    return DARK


def _light_fill(color: str) -> str:
    if color == GREEN:
        return "E8F5EE"
    if color == RED:
        return "FCEDEC"
    return "F1F5F9"


def _example_block_height(report: ReportData) -> int:
    return 6 + len(report.tokens)


def _merge_example_ranges(ws, token_count: int, start_row: int = 1) -> None:
    def merge(start_col: int, start_offset: int, end_col: int, end_offset: int) -> None:
        ws.merge_cells(
            start_row=start_row + start_offset - 1,
            start_column=start_col,
            end_row=start_row + end_offset - 1,
            end_column=end_col,
        )

    for args in [
        (1, 1, 8, 1),
        (1, 2, 2, 2),
        (3, 2, 4, 2),
        (5, 2, 6, 2),
        (7, 2, 8, 2),
        (1, 3, 2, 4),
        (3, 3, 4, 4),
        (5, 3, 6, 4),
        (7, 3, 8, 4),
        (2, 5, 3, 5),
        (4, 5, 5, 5),
        (6, 5, 7, 5),
    ]:
        merge(*args)
    for row in range(start_row + 5, start_row + 5 + token_count):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)


def _style_example_workbook(ws, report: ReportData, start_row: int = 1, freeze_panes: bool = True) -> None:
    token_count = len(report.tokens)
    row = lambda value: start_row + value - 1
    thin = Side(style="thin", color=GRID)
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    no_border = Border()
    fill_summary = PatternFill("solid", fgColor=SUMMARY_BG)
    fill_header = PatternFill("solid", fgColor=HEADER_BG)
    fill_green = PatternFill("solid", fgColor=GREEN_BG)
    fill_red = PatternFill("solid", fgColor=RED_BG)
    fill_bottom = PatternFill("solid", fgColor=BOTTOM_BG)
    fill_white = PatternFill("solid", fgColor=WHITE)

    widths = [23.25, 28.88, 32.88, 28.88, 32.88, 27.25, 36.88, 19.25]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    heights = {1: 54, 2: 24, 3: 18.75, 4: 20.25, 5: 34.5}
    for row_offset, height in heights.items():
        ws.row_dimensions[row(row_offset)].height = height
    for row_idx in range(row(6), row(6) + token_count):
        ws.row_dimensions[row_idx].height = 34.5
    bottom_row = row(6) + token_count
    ws.row_dimensions[bottom_row].height = 13.5

    for row_idx in range(start_row, bottom_row + 1):
        for col in range(1, 9):
            c = ws.cell(row_idx, col)
            c.border = border_all
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = fill_white

    title_cell = ws.cell(row(1), 1)
    title_cell.fill = PatternFill("solid", fgColor=NAVY)
    title_cell.font = Font(name=COMPAT_FONT, size=36, bold=True, color=WHITE)
    title_cell.alignment = Alignment(horizontal="center", vertical="bottom")
    for col in range(2, 9):
        ws.cell(row(1), col).fill = PatternFill("solid", fgColor=NAVY)
        ws.cell(row(1), col).border = no_border

    for col in [1, 3, 5, 7]:
        cell = ws.cell(row(2), col)
        cell.font = Font(name=COMPAT_FONT, size=14, bold=True, color=MUTED)
        cell.fill = fill_summary
        cell.alignment = Alignment(horizontal="center", vertical="bottom")
    for col in [1, 3, 5, 7]:
        cell = ws.cell(row(3), col)
        cell.font = Font(name=COMPAT_FONT, size=17, bold=True, color=DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    summary_fill = fill_green if report.net_usd > 0 else fill_red if report.net_usd < 0 else fill_summary
    summary_color = GREEN if report.net_usd > 0 else RED if report.net_usd < 0 else DARK
    for row_idx in range(row(2), row(5)):
        for col in range(5, 7):
            ws.cell(row_idx, col).fill = summary_fill
            ws.cell(row_idx, col).font = Font(name=COMPAT_FONT, size=17 if row_idx >= row(3) else 14, bold=True, color=summary_color)
    ws.cell(row(2), 5).font = Font(name=COMPAT_FONT, size=14, bold=True, color=summary_color)
    ws.cell(row(3), 5).font = Font(name=COMPAT_FONT, size=17, bold=True, color=summary_color)

    usd_format = '"US$"#,##0.0000;[Red]\\-"US$"#,##0.0000;"US$"0.0000'
    for col in [1, 3, 5]:
        ws.cell(row(3), col).number_format = usd_format
    ws.cell(row(3), 7).number_format = "@"

    for col in [1, 2, 4, 6, 8]:
        cell = ws.cell(row(5), col)
        cell.fill = fill_header
        cell.font = Font(name=COMPAT_FONT, size=14, bold=True, color=DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx in range(row(6), row(6) + token_count):
        symbol = str(ws.cell(row_idx, 1).value or "")
        decimals = _token_decimals(symbol)
        net = Decimal(str(ws.cell(row_idx, 6).value or 0))
        net_fill = fill_green if net > 0 else fill_red if net < 0 else fill_white
        net_color = GREEN if net > 0 else RED if net < 0 else DARK
        ws.cell(row_idx, 1).font = Font(name=COMPAT_FONT, size=14, bold=True, color=DARK)
        ws.cell(row_idx, 1).number_format = "@"
        for col in [2, 4]:
            ws.cell(row_idx, col).font = Font(name=COMPAT_FONT, size=22, color="000000")
            ws.cell(row_idx, col).number_format = _qty_number_format(decimals)
        for col in [6, 7]:
            ws.cell(row_idx, col).fill = net_fill
        ws.cell(row_idx, 6).font = Font(name=COMPAT_FONT, size=22, bold=True, color=net_color)
        ws.cell(row_idx, 6).number_format = _qty_number_format(decimals)
        ws.cell(row_idx, 8).font = Font(name=COMPAT_FONT, size=14, color=DARK)
        ws.cell(row_idx, 8).number_format = "@"

    for col in range(1, 9):
        ws.cell(bottom_row, col).fill = fill_bottom
        ws.cell(bottom_row, col).font = Font(name=COMPAT_FONT, size=14, bold=True, color="123B2A")
    if freeze_panes:
        ws.freeze_panes = f"A{row(6)}"


def _draw_cell(
    draw: ImageDraw.ImageDraw,
    xs: list[int],
    ys: list[int],
    start_col: int,
    start_row: int,
    end_col: int,
    end_row: int,
    fill: str,
) -> None:
    draw.rectangle(
        (xs[start_col], ys[start_row], xs[end_col + 1], ys[end_row + 1]),
        fill="#" + fill,
        outline="#" + GRID,
        width=1,
    )


def _draw_grid(draw: ImageDraw.ImageDraw, xs: list[int], ys: list[int], color: str) -> None:
    for x in xs:
        draw.line((x, ys[0], x, ys[-1]), fill="#" + color, width=1)
    for y in ys:
        draw.line((xs[0], y, xs[-1], y), fill="#" + color, width=1)


def _draw_center(
    draw: ImageDraw.ImageDraw,
    box: tuple[int, int, int, int],
    text: str,
    font: ImageFont.FreeTypeFont,
    color: str,
    y_offset: int = 0,
) -> None:
    bbox = draw.textbbox((0, 0), text, font=font)
    tw = bbox[2] - bbox[0]
    th = bbox[3] - bbox[1]
    x = box[0] + ((box[2] - box[0]) - tw) / 2
    y = box[1] + ((box[3] - box[1]) - th) / 2 - bbox[1] + y_offset
    draw.text((x, y), text, fill="#" + color, font=font)


def _money4(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP)


def _fmt_usd4(value: Decimal) -> str:
    q = _money4(value)
    sign = "-" if q < 0 else ""
    q = abs(q)
    return f'{sign}US${q:,.4f}'


def _token_decimals(symbol: str) -> int:
    return TOKEN_DECIMALS.get(symbol.upper(), 1)


def _round_token_qty(value: Decimal, symbol: str) -> Decimal:
    decimals = _token_decimals(symbol)
    quant = Decimal("1").scaleb(-decimals)
    return value.quantize(quant, rounding=ROUND_HALF_UP)


def _fmt_token_qty(value: Decimal, symbol: str) -> str:
    decimals = _token_decimals(symbol)
    return f"{_round_token_qty(value, symbol):,.{decimals}f}"


def _qty_number_format(decimals: int) -> str:
    if decimals <= 0:
        return '#,##0'
    return '#,##0.' + ('0' * decimals)


def _status_label(value: Decimal) -> str:
    if value > 0:
        return "净流入"
    if value < 0:
        return "净流出"
    return "持平"


def _font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont:
    candidates = [
        "/System/Library/Fonts/PingFang.ttc",
        "/System/Library/Fonts/STHeiti Light.ttc",
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        "/Library/Fonts/Arial Unicode.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]
    for candidate in candidates:
        try:
            return ImageFont.truetype(candidate, size=size)
        except OSError:
            continue
    return ImageFont.load_default()
