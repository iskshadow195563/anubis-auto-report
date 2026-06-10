from __future__ import annotations

import tempfile
import unittest
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from auto_report.parser import parse_report_text
from auto_report.renderer import GREEN_BG, RED_BG, render_report_files, workbook_has_formulas


SAMPLE_TEXT = """
今日转入
US$580,860.03
币种    数量    美金计价
A
3,842.950295    US$295.14
BNB
6.568761    US$4,633.93
DAI
570,098.061137    US$570,098.06
ETH
0.250791    US$502.51
LGNS
0.000000    US$0.00
POL
47.000000    US$4.36
SLGNS
0.000000    US$0.00
USDC
0.000000    US$0.00
USDT
5,326.029608    US$5,326.03
今日转出
US$561,225.66
币种    数量    美金计价
A
0.000000    US$0.00
BNB
0.993514    US$700.87
DAI
555,824.915123    US$555,824.92
ETH
0.000000    US$0.00
LGNS
0.000000    US$0.00
POL
48,046.814106    US$4,453.94
SLGNS
0.000000    US$0.00
USDC
0.000000    US$0.00
USDT
245.930000    US$245.93
"""


class ParserRendererTest(unittest.TestCase):
    def test_parse_sample_text(self) -> None:
        report = parse_report_text(SAMPLE_TEXT)
        self.assertEqual(report.total_in_usd, Decimal("580860.03"))
        self.assertEqual(report.total_out_usd, Decimal("561225.66"))
        self.assertEqual(report.status, "净流入")
        self.assertEqual(len(report.tokens), 9)
        pol = next(item for item in report.tokens if item.symbol == "POL")
        self.assertEqual(pol.status, "净流出")

    def test_rendered_xlsx_has_no_formula(self) -> None:
        report = parse_report_text(SAMPLE_TEXT)
        with tempfile.TemporaryDirectory() as tmp:
            image_path, xlsx_path = render_report_files(report, Path(tmp))
            self.assertTrue(image_path.exists())
            self.assertTrue(xlsx_path.exists())
            self.assertFalse(workbook_has_formulas(xlsx_path))

    def test_xlsx_uses_report_then_history_sheets(self) -> None:
        report = parse_report_text(SAMPLE_TEXT)
        with tempfile.TemporaryDirectory() as tmp:
            _, xlsx_path = render_report_files(report, Path(tmp))
            wb = load_workbook(xlsx_path, data_only=False)
            self.assertEqual(wb.sheetnames[0], f"{report.captured_at.month}月{report.captured_at.day}汇总")
            self.assertEqual(wb.sheetnames[1], "歷史記錄")
            self.assertNotIn("原生兼容", wb.sheetnames)
            ws = wb[wb.sheetnames[0]]
            self.assertEqual(ws["A1"].value, f"币种资金流日报 - {report.captured_at.strftime('%Y-%m-%d')}")
            self.assertEqual(ws["A3"].value, float(Decimal("580860.0300")))
            self.assertEqual(ws["C3"].value, float(Decimal("561225.6600")))
            self.assertEqual(ws["E3"].value, float(Decimal("19634.3700")))
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        self.assertFalse(isinstance(cell.value, str) and cell.value.startswith("="))

    def test_history_sheet_stacks_latest_daily_report_first(self) -> None:
        hkt = timezone(timedelta(hours=8))
        old_report = parse_report_text(SAMPLE_TEXT, captured_at=datetime(2026, 5, 17, 23, 30, tzinfo=hkt))
        new_report = parse_report_text(SAMPLE_TEXT, captured_at=datetime(2026, 6, 3, 23, 30, tzinfo=hkt))
        with tempfile.TemporaryDirectory() as tmp:
            _, xlsx_path = render_report_files(new_report, Path(tmp), history_reports=[new_report, old_report])
            wb = load_workbook(xlsx_path, data_only=False)
            self.assertEqual(wb.sheetnames[1], "歷史記錄")
            ws = wb["歷史記錄"]
            self.assertEqual(ws["A1"].value, "币种资金流日报 - 2026-06-03")
            self.assertEqual(ws["A16"].value, "币种资金流日报 - 2026-05-17")
            self.assertTrue(str(ws["E3"].fill.fgColor.rgb).endswith(GREEN_BG))
            self.assertTrue(str(ws["F11"].fill.fgColor.rgb).endswith(RED_BG))
            self.assertFalse(workbook_has_formulas(xlsx_path))


if __name__ == "__main__":
    unittest.main()
